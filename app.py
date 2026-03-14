import sqlite3
import json
import requests
import os
import subprocess
import sys
import io
from datetime import datetime, timedelta
from flask import Flask, render_template, jsonify, request, Response, stream_with_context, session, redirect, url_for, send_file
import queue
import threading
import functools

from dotenv import load_dotenv
load_dotenv()
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
DB_PATH = os.environ.get("DB_PATH", os.path.join(os.path.dirname(os.path.abspath(__file__)), "orders.db"))

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-in-prod")


def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# Инициализируем БД при старте (нужно и под gunicorn)
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, name TEXT, phone TEXT, service TEXT,
            address TEXT, date TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            reminder_sent INTEGER DEFAULT 0,
            admin_message_id INTEGER,
            rating INTEGER DEFAULT NULL,
            review_sent INTEGER DEFAULT 0,
            status TEXT DEFAULT 'new'
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, user_name TEXT, direction TEXT,
            text TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS workers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            telegram_id INTEGER
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS referrals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            referrer_id INTEGER NOT NULL,
            referred_id INTEGER NOT NULL,
            notified INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(referred_id)
        )
    """)
    for col in ["admin_message_id INTEGER", "rating INTEGER DEFAULT NULL",
                "review_sent INTEGER DEFAULT 0", "status TEXT DEFAULT 'new'",
                "price REAL DEFAULT NULL", "executor TEXT DEFAULT NULL",
                "executor_id INTEGER DEFAULT NULL"]:
        try:
            c.execute(f"ALTER TABLE orders ADD COLUMN {col}")
        except Exception:
            pass
    conn.commit()
    conn.close()

init_db()

# Запускаем бота как фоновый процесс с авто-перезапуском
def _start_bot():
    import time
    bot_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bot.py")
    while True:
        subprocess.run([sys.executable, bot_path])
        time.sleep(10)  # пауза перед перезапуском (даёт время старому контейнеру остановиться)

threading.Thread(target=_start_bot, daemon=True).start()

# SSE очередь для уведомлений
sse_clients = []
sse_lock = threading.Lock()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, name TEXT, phone TEXT, service TEXT,
            address TEXT, date TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            reminder_sent INTEGER DEFAULT 0,
            admin_message_id INTEGER,
            rating INTEGER DEFAULT NULL,
            review_sent INTEGER DEFAULT 0,
            status TEXT DEFAULT 'new'
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, user_name TEXT, direction TEXT,
            text TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS workers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            telegram_id INTEGER
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS referrals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            referrer_id INTEGER NOT NULL,
            referred_id INTEGER NOT NULL,
            notified INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(referred_id)
        )
    """)
    for col in ["admin_message_id INTEGER", "rating INTEGER DEFAULT NULL",
                "review_sent INTEGER DEFAULT 0", "status TEXT DEFAULT 'new'",
                "price REAL DEFAULT NULL", "executor TEXT DEFAULT NULL",
                "executor_id INTEGER DEFAULT NULL"]:
        try:
            c.execute(f"ALTER TABLE orders ADD COLUMN {col}")
        except Exception:
            pass
    conn.commit()
    conn.close()


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def push_sse_event(data):
    with sse_lock:
        dead = []
        for q in sse_clients:
            try:
                q.put_nowait(data)
            except Exception:
                dead.append(q)
        for q in dead:
            sse_clients.remove(q)


# ─── Авторизация ──────────────────────────────────────────────────────────────


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        crm_password = os.environ.get("CRM_PASS") or os.environ.get("CRM_PASSWORD") or "vid2026"
        if request.form.get("password") == crm_password:
            session["logged_in"] = True
            return redirect(url_for("dashboard"))
        error = "Неверный пароль"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─── Страницы ─────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html")


@app.route("/chats")
@login_required
def chats():
    return render_template("chats.html")


@app.route("/kanban")
@login_required
def kanban():
    return render_template("kanban.html")


@app.route("/schedule")
@login_required
def schedule():
    return render_template("schedule.html")


# ─── API: Статистика ──────────────────────────────────────────────────────────

@app.route("/api/stats")
@login_required
def api_stats():
    conn = get_db()
    c = conn.cursor()

    today = datetime.now().strftime("%Y-%m-%d")
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    month_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

    # Заявки сегодня
    c.execute("SELECT COUNT(*) FROM orders WHERE created_at >= ?", (today,))
    today_count = c.fetchone()[0]

    # Заявки за неделю
    c.execute("SELECT COUNT(*) FROM orders WHERE created_at >= ?", (week_ago,))
    week_count = c.fetchone()[0]

    # Всего заявок
    c.execute("SELECT COUNT(*) FROM orders")
    total_count = c.fetchone()[0]

    # Уникальные клиенты
    c.execute("SELECT COUNT(DISTINCT user_id) FROM orders")
    clients_count = c.fetchone()[0]

    # Средний рейтинг
    c.execute("SELECT AVG(rating) FROM orders WHERE rating IS NOT NULL")
    avg_rating = c.fetchone()[0]
    avg_rating = round(avg_rating, 1) if avg_rating else 0

    # Выручка сегодня
    c.execute("SELECT COALESCE(SUM(price), 0) FROM orders WHERE created_at >= ? AND price IS NOT NULL", (today,))
    today_revenue = int(c.fetchone()[0])

    # Выручка за месяц
    c.execute("SELECT COALESCE(SUM(price), 0) FROM orders WHERE created_at >= ? AND price IS NOT NULL", (month_ago,))
    month_revenue = int(c.fetchone()[0])

    # Выручка по дням (последние 14 дней)
    c.execute("""
        SELECT DATE(created_at) as day, COALESCE(SUM(price), 0) as revenue
        FROM orders
        WHERE created_at >= DATE('now', '-14 days')
        GROUP BY DATE(created_at)
        ORDER BY day
    """)
    revenue_data = [{"day": r["day"], "revenue": r["revenue"]} for r in c.fetchall()]

    # Заявки по дням (последние 14 дней)
    c.execute("""
        SELECT DATE(created_at) as day, COUNT(*) as cnt
        FROM orders
        WHERE created_at >= DATE('now', '-14 days')
        GROUP BY DATE(created_at)
        ORDER BY day
    """)
    days_data = [{"day": r["day"], "count": r["cnt"]} for r in c.fetchall()]

    # Заявки по услугам
    c.execute("""
        SELECT service, COUNT(*) as cnt
        FROM orders
        GROUP BY service
        ORDER BY cnt DESC
        LIMIT 8
    """)
    services_data = [{"service": r["service"] or "Не указано", "count": r["cnt"]} for r in c.fetchall()]

    # По статусам
    c.execute("""
        SELECT status, COUNT(*) as cnt FROM orders GROUP BY status
    """)
    statuses = {r["status"]: r["cnt"] for r in c.fetchall()}

    # Последние 10 заявок
    c.execute("""
        SELECT id, name, phone, service, address, date, status, rating, created_at
        FROM orders ORDER BY id DESC LIMIT 10
    """)
    recent = []
    for r in c.fetchall():
        recent.append({
            "id": r["id"],
            "name": r["name"],
            "phone": r["phone"],
            "service": r["service"],
            "address": r["address"],
            "date": r["date"],
            "status": r["status"] or "new",
            "rating": r["rating"],
            "created_at": r["created_at"],
        })

    # Воронка конверсии
    c.execute("SELECT COUNT(DISTINCT user_id) FROM messages WHERE direction='in'")
    funnel_leads = c.fetchone()[0]
    funnel_orders = total_count
    funnel_done = statuses.get("done", 0)
    funnel_cancelled = statuses.get("cancelled", 0)
    funnel = {
        "leads": funnel_leads,
        "orders": funnel_orders,
        "done": funnel_done,
        "cancelled": funnel_cancelled,
        "conv_order": round(funnel_orders / funnel_leads * 100, 1) if funnel_leads else 0,
        "conv_done": round(funnel_done / funnel_orders * 100, 1) if funnel_orders else 0,
    }

    # Средний чек (по заявкам с ценой)
    c.execute("SELECT AVG(price) FROM orders WHERE price IS NOT NULL AND price > 0")
    avg_check = c.fetchone()[0]
    avg_check = int(avg_check) if avg_check else 0

    # Топ клиентов по LTV (суммарная выручка)
    c.execute("""
        SELECT name,
               COUNT(*) as orders_count,
               COALESCE(SUM(price), 0) as ltv,
               COALESCE(AVG(price), 0) as avg_price
        FROM orders
        WHERE name IS NOT NULL AND name != ''
        GROUP BY user_id
        ORDER BY ltv DESC
        LIMIT 10
    """)
    top_clients = [
        {
            "name": r["name"],
            "orders": r["orders_count"],
            "ltv": int(r["ltv"]),
            "avg_check": int(r["avg_price"]),
        }
        for r in c.fetchall()
    ]

    conn.close()
    return jsonify({
        "today_count": today_count,
        "week_count": week_count,
        "total_count": total_count,
        "clients_count": clients_count,
        "avg_rating": avg_rating,
        "today_revenue": today_revenue,
        "month_revenue": month_revenue,
        "revenue_data": revenue_data,
        "days_data": days_data,
        "services_data": services_data,
        "statuses": statuses,
        "recent": recent,
        "funnel": funnel,
        "avg_check": avg_check,
        "top_clients": top_clients,
    })


# ─── API: Чаты ────────────────────────────────────────────────────────────────

@app.route("/api/clients")
@login_required
def api_clients():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT m.user_id,
               m.user_name,
               MAX(m.timestamp) as last_time,
               (SELECT text FROM messages m2 WHERE m2.user_id = m.user_id ORDER BY m2.id DESC LIMIT 1) as last_msg,
               COUNT(CASE WHEN m.direction='in' THEN 1 END) as in_count,
               CAST(strftime('%H', (
                   SELECT timestamp FROM messages m3
                   WHERE m3.user_id = m.user_id AND m3.direction='in'
                   ORDER BY m3.id DESC LIMIT 1
               )) AS INTEGER) as last_in_hour
        FROM messages m
        GROUP BY m.user_id
        ORDER BY last_time DESC
    """)
    rows = c.fetchall()
    c.execute("SELECT referrer_id, COUNT(*) as cnt FROM referrals GROUP BY referrer_id")
    ref_counts = {r["referrer_id"]: r["cnt"] for r in c.fetchall()}
    clients = []
    for r in rows:
        h = r["last_in_hour"]
        night_lead = h is not None and (h >= 20 or h < 9)
        clients.append({
            "user_id": r["user_id"],
            "user_name": r["user_name"],
            "last_time": r["last_time"],
            "last_msg": r["last_msg"],
            "in_count": r["in_count"],
            "night_lead": night_lead,
            "referrals": ref_counts.get(r["user_id"], 0),
        })
    conn.close()
    return jsonify(clients)


@app.route("/api/messages/<int:user_id>")
@login_required
def api_messages(user_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, user_id, user_name, direction, text, timestamp
        FROM messages WHERE user_id=?
        ORDER BY id ASC
    """, (user_id,))
    msgs = []
    for r in c.fetchall():
        msgs.append({
            "id": r["id"],
            "user_id": r["user_id"],
            "user_name": r["user_name"],
            "direction": r["direction"],
            "text": r["text"],
            "timestamp": r["timestamp"],
        })
    conn.close()
    return jsonify(msgs)


@app.route("/api/send/<int:user_id>", methods=["POST"])
@login_required
def api_send(user_id):
    data = request.get_json()
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"ok": False, "error": "Пустое сообщение"}), 400

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    resp = requests.post(url, json={"chat_id": user_id, "text": text})
    if not resp.ok:
        return jsonify({"ok": False, "error": resp.text}), 500

    # сохранить в БД
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO messages (user_id, user_name, direction, text) VALUES (?, ?, ?, ?)",
        (user_id, "ООО ВИД", "out", text)
    )
    conn.commit()
    msg_id = c.lastrowid
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.close()

    push_sse_event(json.dumps({
        "type": "message",
        "user_id": user_id,
        "direction": "out",
        "text": text,
        "timestamp": ts,
    }))

    return jsonify({"ok": True, "id": msg_id, "timestamp": ts})


# ─── API: Канбан ──────────────────────────────────────────────────────────────

@app.route("/api/orders")
@login_required
def api_orders():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, user_id, name, phone, service, address, date, status, rating, created_at, executor
        FROM orders ORDER BY id DESC
    """)
    orders = []
    for r in c.fetchall():
        orders.append({
            "id": r["id"],
            "user_id": r["user_id"],
            "name": r["name"],
            "phone": r["phone"],
            "service": r["service"],
            "address": r["address"],
            "date": r["date"],
            "status": r["status"] or "new",
            "rating": r["rating"],
            "created_at": r["created_at"],
            "executor": r["executor"],
        })
    conn.close()
    return jsonify(orders)


@app.route("/api/orders/<int:order_id>/status", methods=["PATCH"])
@login_required
def api_update_status(order_id):
    data = request.get_json()
    new_status = data.get("status")
    allowed = ("new", "in_progress", "done", "cancelled")
    if new_status not in allowed:
        return jsonify({"ok": False, "error": "Неверный статус"}), 400
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE orders SET status=? WHERE id=?", (new_status, order_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─── API: Фото ────────────────────────────────────────────────────────────────

@app.route("/api/photo/<file_id>")
@login_required
def api_photo(file_id):
    info = requests.get(
        f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getFile",
        params={"file_id": file_id}, timeout=10
    )
    if not info.ok:
        return "Not found", 404
    file_path = info.json().get("result", {}).get("file_path")
    if not file_path:
        return "Not found", 404
    img = requests.get(
        f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}", timeout=15
    )
    if not img.ok:
        return "Not found", 404
    content_type = img.headers.get("Content-Type", "image/jpeg")
    return Response(img.content, mimetype=content_type)


# ─── API: Расписание ──────────────────────────────────────────────────────────

@app.route("/api/schedule")
@login_required
def api_schedule():
    conn = get_db()
    c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    # Все активные заявки (new + in_progress) + выполненные созданные сегодня
    c.execute("""
        SELECT id, user_id, name, phone, service, address, date, status,
               rating, created_at, executor, price
        FROM orders
        WHERE status IN ('new', 'in_progress')
           OR (status = 'done' AND created_at >= ?)
        ORDER BY
            CASE status WHEN 'in_progress' THEN 0 WHEN 'new' THEN 1 ELSE 2 END,
            created_at ASC
    """, (today,))
    orders = []
    for r in c.fetchall():
        orders.append({
            "id": r["id"], "name": r["name"], "phone": r["phone"],
            "service": r["service"], "address": r["address"], "date": r["date"],
            "status": r["status"] or "new", "rating": r["rating"],
            "created_at": r["created_at"], "executor": r["executor"],
            "price": r["price"],
        })
    # Статистика дня
    c.execute("SELECT COUNT(*) FROM orders WHERE created_at >= ?", (today,))
    created_today = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM orders WHERE status='in_progress'")
    in_progress = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM orders WHERE status='done' AND created_at >= ?", (today,))
    done_today = c.fetchone()[0]
    conn.close()
    return jsonify({
        "orders": orders,
        "stats": {"created_today": created_today, "in_progress": in_progress, "done_today": done_today},
    })


# ─── API: Сотрудники ──────────────────────────────────────────────────────────

@app.route("/api/workers")
@login_required
def api_workers():
    conn = get_db()
    rows = conn.execute("SELECT id, name, telegram_id FROM workers ORDER BY name").fetchall()
    conn.close()
    return jsonify([{"id": r["id"], "name": r["name"], "telegram_id": r["telegram_id"]} for r in rows])


@app.route("/api/workers", methods=["POST"])
@login_required
def api_add_worker():
    data = request.get_json()
    name = data.get("name", "").strip()
    telegram_id = data.get("telegram_id") or None
    if not name:
        return jsonify({"ok": False, "error": "Имя обязательно"}), 400
    conn = get_db()
    conn.execute("INSERT INTO workers (name, telegram_id) VALUES (?, ?)", (name, telegram_id))
    conn.commit()
    worker_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "id": worker_id})


@app.route("/api/workers/<int:worker_id>", methods=["DELETE"])
@login_required
def api_delete_worker(worker_id):
    conn = get_db()
    conn.execute("DELETE FROM workers WHERE id=?", (worker_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/orders/<int:order_id>/executor", methods=["PATCH"])
@login_required
def api_assign_executor(order_id):
    data = request.get_json()
    worker_id = data.get("worker_id")

    executor_name = None
    executor_tg_id = None
    if worker_id:
        conn = get_db()
        w = conn.execute("SELECT name, telegram_id FROM workers WHERE id=?", (worker_id,)).fetchone()
        conn.close()
        if not w:
            return jsonify({"ok": False, "error": "Сотрудник не найден"}), 404
        executor_name = w["name"]
        executor_tg_id = w["telegram_id"]

    conn = get_db()
    conn.execute("UPDATE orders SET executor=?, executor_id=? WHERE id=?",
                 (executor_name, executor_tg_id, order_id))
    conn.commit()
    order = conn.execute(
        "SELECT name, service, address, date FROM orders WHERE id=?", (order_id,)
    ).fetchone()
    conn.close()

    if executor_tg_id and order:
        text = (
            f"📋 Вам назначена заявка #{order_id}\n\n"
            f"👤 Клиент: {order['name']}\n"
            f"🛠 Услуга: {order['service']}\n"
            f"🏠 Адрес: {order['address']}\n"
            f"📅 Дата: {order['date']}\n\n"
            "Удачи в работе! — ООО ВИД 🏢"
        )
        try:
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                json={"chat_id": executor_tg_id, "text": text}
            )
        except Exception:
            pass

    return jsonify({"ok": True, "executor": executor_name})


# ─── API: Экспорт Excel ───────────────────────────────────────────────────────

@app.route("/api/export")
@login_required
def api_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, name, phone, service, address, date, status, price, rating, created_at
        FROM orders ORDER BY id DESC
    """)
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки ООО ВИД"

    headers = ["№", "Клиент", "Телефон", "Услуга", "Адрес", "Дата", "Статус", "Выручка (₽)", "Оценка", "Создана"]
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    STATUS_RU = {"new": "Новая", "in_progress": "В работе", "done": "Выполнена", "cancelled": "Отменена"}
    for row in rows:
        ws.append([
            row["id"],
            row["name"] or "",
            row["phone"] or "",
            row["service"] or "",
            row["address"] or "",
            row["date"] or "",
            STATUS_RU.get(row["status"] or "new", row["status"] or ""),
            row["price"] or "",
            row["rating"] or "",
            row["created_at"] or "",
        ])

    col_widths = [6, 22, 16, 30, 30, 18, 12, 14, 8, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"vid_orders_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── API: Рассылка ────────────────────────────────────────────────────────────

@app.route("/api/broadcast/count")
@login_required
def api_broadcast_count():
    active_only = request.args.get("active_only", "true") == "true"
    conn = get_db()
    c = conn.cursor()
    if active_only:
        cutoff = (datetime.now() - timedelta(days=180)).strftime("%Y-%m-%d")
        c.execute("SELECT COUNT(DISTINCT user_id) FROM messages WHERE timestamp >= ?", (cutoff,))
    else:
        c.execute("SELECT COUNT(DISTINCT user_id) FROM messages")
    count = c.fetchone()[0]
    conn.close()
    return jsonify({"count": count})


@app.route("/api/broadcast", methods=["POST"])
@login_required
def api_broadcast():
    data = request.get_json()
    text = data.get("text", "").strip()
    active_only = data.get("active_only", True)

    if not text:
        return jsonify({"ok": False, "error": "Пустое сообщение"}), 400

    conn = get_db()
    c = conn.cursor()
    if active_only:
        cutoff = (datetime.now() - timedelta(days=180)).strftime("%Y-%m-%d")
        c.execute("SELECT DISTINCT user_id FROM messages WHERE timestamp >= ?", (cutoff,))
    else:
        c.execute("SELECT DISTINCT user_id FROM messages")
    user_ids = [r["user_id"] for r in c.fetchall()]
    conn.close()

    sent = 0
    failed = 0
    tg_url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for uid in user_ids:
        resp = requests.post(tg_url, json={"chat_id": uid, "text": text})
        if resp.ok:
            sent += 1
            conn = get_db()
            conn.execute(
                "INSERT INTO messages (user_id, user_name, direction, text, timestamp) VALUES (?, ?, ?, ?, ?)",
                (uid, "ООО ВИД", "out", text, ts)
            )
            conn.commit()
            conn.close()
            push_sse_event(json.dumps({
                "type": "message", "user_id": uid,
                "direction": "out", "text": text, "timestamp": ts,
            }))
        else:
            failed += 1

    return jsonify({"ok": True, "sent": sent, "failed": failed, "total": len(user_ids)})


# ─── API: Последние заявки (пагинация) ───────────────────────────────────────

@app.route("/api/recent")
@login_required
def api_recent():
    per_page = 10
    page = max(1, int(request.args.get("page", 1)))
    offset = (page - 1) * per_page
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM orders")
    total = c.fetchone()[0]
    c.execute("""
        SELECT id, name, phone, service, address, date, status, rating
        FROM orders ORDER BY id DESC LIMIT ? OFFSET ?
    """, (per_page, offset))
    orders = []
    for r in c.fetchall():
        orders.append({
            "id": r["id"], "name": r["name"], "phone": r["phone"],
            "service": r["service"], "address": r["address"],
            "date": r["date"], "status": r["status"] or "new",
            "rating": r["rating"],
        })
    conn.close()
    return jsonify({"orders": orders, "total": total, "page": page, "per_page": per_page})


# ─── API: Статус бота ────────────────────────────────────────────────────────

@app.route("/api/bot/status")
@login_required
def api_bot_status():
    try:
        resp = requests.get(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getMe",
            timeout=5
        )
        if resp.ok and resp.json().get("ok"):
            return jsonify({"ok": True})
    except Exception:
        pass
    return jsonify({"ok": False})


# ─── API: Карточка клиента ────────────────────────────────────────────────────

@app.route("/api/client/<int:user_id>")
@login_required
def api_client(user_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT name, phone, COUNT(*) as orders_count,
               COALESCE(SUM(price), 0) as ltv,
               COALESCE(AVG(price), 0) as avg_check
        FROM orders WHERE user_id=?
    """, (user_id,))
    row = c.fetchone()
    c.execute("""
        SELECT id, service, address, date, status, price, rating, created_at
        FROM orders WHERE user_id=? ORDER BY id DESC LIMIT 20
    """, (user_id,))
    orders = [dict(r) for r in c.fetchall()]
    c.execute("SELECT COUNT(*) FROM referrals WHERE referrer_id=?", (user_id,))
    referrals = c.fetchone()[0]
    conn.close()
    return jsonify({
        "user_id": user_id,
        "name": row["name"] if row else None,
        "phone": row["phone"] if row else None,
        "orders_count": row["orders_count"] if row else 0,
        "ltv": int(row["ltv"]) if row else 0,
        "avg_check": int(row["avg_check"]) if row else 0,
        "referrals": referrals,
        "orders": orders,
    })


# ─── SSE: Уведомления ─────────────────────────────────────────────────────────

@app.route("/api/events")
@login_required
def api_events():
    q = queue.Queue(maxsize=50)
    with sse_lock:
        sse_clients.append(q)

    def generate():
        try:
            yield "data: {\"type\": \"connected\"}\n\n"
            while True:
                try:
                    msg = q.get(timeout=20)
                    yield f"data: {msg}\n\n"
                except queue.Empty:
                    yield ": ping\n\n"
        except GeneratorExit:
            with sse_lock:
                if q in sse_clients:
                    sse_clients.remove(q)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        }
    )


if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
